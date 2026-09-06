"""Tests for searching inside an attachment.

The point of search is that it scans the WHOLE file, so the fixtures here are
deliberately longer than any per-read page: a test that passes on ten rows would
not have caught the row cap that made this feature necessary.
"""
import io
import os
import sys

import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import attachment_readers as ar


def _xlsx(rows, sheet_title="Sheet1"):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _big_branch_file(total=1200):
    """A file whose interesting rows sit well past any old truncation point."""
    rows = [("סניף", "עיר", "אחראי", "סטטוס")]
    for i in range(total):
        rows.append((f"סניף {i}", "חיפה", "דנה", "בוצע"))
    rows.append(("סניף 900", "רמלה", "איתי", "בוצע"))
    rows.append(("סניף 901", 'ראשל"צ', "ניקיטה", "ממתין"))
    rows.append(("סניף 902", "קרית אונו", "איתי", "ממתין"))
    rows.append(("סניף 903", "Rishon LeZion", "Nikita", "done"))
    rows.append(("סניף 904", "לוד", "דנה", "בוצע"))
    return _xlsx(rows)


# --- normalisation -----------------------------------------------------------

@pytest.mark.parametrize("a,b", [
    ("קריית אונו", "קרית אונו"),      # כתיב מלא vs חסר
    ('ראשל"צ', "ראשל״צ"),             # ASCII quote vs Hebrew gershayim
    ("אור-יהודה", "אור יהודה"),        # maqaf vs space
    ("Rishon LeZion", "rishon lezion"),
])
def test_spellings_of_the_same_place_normalise_together(a, b):
    assert ar._normalise(a) == ar._normalise(b)


def test_a_hebrew_city_expands_to_its_english_spelling():
    expanded = ar.expand_keyword("ראשון לציון")
    assert ar._normalise("rishon lezion") in expanded
    assert ar._normalise('ראשל"צ') in expanded


def test_an_unknown_term_expands_to_just_itself():
    assert ar.expand_keyword("ניקיטה") == [ar._normalise("ניקיטה")]


def test_every_city_itai_listed_is_known():
    for city in ["ראשון לציון", "רמלה", "לוד", "קריית אונו",
                 "קריית עקרון", "יבנה", "אור יהודה"]:
        assert len(ar.expand_keyword(city)) > 1, f"{city} has no aliases"


# --- matching ----------------------------------------------------------------

def test_a_short_city_does_not_match_inside_a_longer_word():
    # The reason matching is on whole tokens: 'לוד' and 'Lod' are short enough
    # that substring matching would produce confident nonsense.
    assert not ar._matches(ar._tokens("Lodging expenses"), ["lod"])
    assert ar._matches(ar._tokens("סניף לוד מרכז"), ["לוד"])


def test_a_two_word_city_matches_inside_a_longer_cell():
    assert ar._matches(ar._tokens("סניף אור יהודה - קניון"), ar._normalise("אור יהודה").split())


# --- search ------------------------------------------------------------------

def test_search_scans_rows_far_past_any_read_page():
    raw = _big_branch_file()
    result = ar.search("branches.xlsx", raw, "רמלה")
    assert "רמלה" in result and "איתי" in result
    assert "1,206" in result   # 1 header + 1200 filler + 5 target rows


def test_search_expands_a_hebrew_city_to_its_abbreviation_and_english():
    raw = _big_branch_file()
    result = ar.search("branches.xlsx", raw, "ראשון לציון")
    assert "סניף 901" in result   # written ראשל"צ
    assert "סניף 903" in result   # written Rishon LeZion


def test_search_finds_a_city_spelled_with_the_other_orthography():
    raw = _big_branch_file()
    result = ar.search("branches.xlsx", raw, "קריית אונו")
    assert "סניף 902" in result


def test_cross_reference_keeps_only_rows_naming_one_of_the_people():
    raw = _big_branch_file()
    result = ar.search(
        "branches.xlsx", raw,
        "ראשון לציון, רמלה, לוד, קריית אונו, קריית עקרון, יבנה, אור יהודה",
        must_also_match="איתי, ניקיטה",
    )
    assert "סניף 900" in result   # רמלה + איתי
    assert "סניף 901" in result   # ראשל"צ + ניקיטה
    assert "סניף 902" in result   # קרית אונו + איתי
    # לוד is in the territory but its row belongs to דנה, so the AND drops it.
    assert "סניף 904" not in result


def test_cross_reference_matches_a_name_written_in_english():
    raw = _big_branch_file()
    result = ar.search("branches.xlsx", raw, "ראשון לציון", must_also_match="Nikita")
    assert "סניף 903" in result
    assert "סניף 901" not in result


def test_a_row_outside_the_territory_is_never_returned():
    raw = _big_branch_file()
    result = ar.search("branches.xlsx", raw, "רמלה, לוד")
    assert "חיפה" not in result


def test_no_match_says_how_much_was_scanned_rather_than_failing():
    raw = _big_branch_file()
    result = ar.search("branches.xlsx", raw, "אילת")
    assert "לא נמצאה" in result
    assert "1,206" in result


def test_search_reports_the_sheet_a_hit_came_from():
    raw = _xlsx([("עיר", "אחראי"), ("יבנה", "איתי")], sheet_title="מרץ")
    result = ar.search("f.xlsx", raw, "יבנה")
    assert "מרץ" in result


def test_search_works_on_a_semicolon_separated_hebrew_csv():
    raw = "עיר;אחראי\nיבנה;איתי\nחיפה;דנה\n".encode("utf-8")
    result = ar.search("f.csv", raw, "יבנה", must_also_match="איתי")
    assert "יבנה" in result and "חיפה" not in result


def test_search_refuses_an_unreadable_type_with_an_explanation():
    result = ar.search("scan.jpg", b"\xff\xd8\xff", "רמלה")
    assert result.startswith("❌") and "תמונה" in result


def test_search_without_keywords_says_so():
    assert "מילות חיפוש" in ar.search("f.xlsx", _xlsx([("a",)]), "")


def test_search_caps_what_it_prints_but_still_counts_every_hit():
    rows = [("עיר", "אחראי")] + [("רמלה", "איתי")] * (ar.MAX_MATCHES_RETURNED + 50)
    result = ar.search("f.xlsx", _xlsx(rows), "רמלה")
    assert "כדאי לצמצם" in result


# --- paging ------------------------------------------------------------------

def test_a_long_file_is_paged_rather_than_cut_off():
    raw = _big_branch_file()
    first = ar.extract_text("branches.xlsx", raw)
    assert "חלק 1 מתוך" in first
    assert "אפשר לבקש את חלק 2" in first


def test_a_later_part_returns_different_content_and_the_last_says_so():
    raw = _big_branch_file()
    first = ar.extract_text("branches.xlsx", raw, part=1)
    total = int(first.split("מתוך ")[1].split("]")[0])
    assert total > 1
    second = ar.extract_text("branches.xlsx", raw, part=2)
    assert second != first
    last = ar.extract_text("branches.xlsx", raw, part=total)
    assert "הקובץ נקרא במלואו" in last


def test_the_last_page_contains_the_rows_the_old_cap_hid():
    raw = _big_branch_file()
    total = int(ar.extract_text("branches.xlsx", raw).split("מתוך ")[1].split("]")[0])
    last = ar.extract_text("branches.xlsx", raw, part=total)
    assert "סניף 904" in last


def test_asking_past_the_end_returns_the_last_page_instead_of_nothing():
    raw = _big_branch_file()
    assert "הקובץ נקרא במלואו" in ar.extract_text("branches.xlsx", raw, part=9999)


def test_a_short_file_is_not_paged_at_all():
    raw = _xlsx([("עיר", "אחראי"), ("יבנה", "איתי")])
    assert "חלק" not in ar.extract_text("f.xlsx", raw)


def test_iter_rows_yields_every_row_with_no_limit():
    raw = _big_branch_file(total=900)
    assert len(list(ar.iter_rows("f.xlsx", raw))) == 906
